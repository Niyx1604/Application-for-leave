#!/usr/bin/env python3
"""
Leave Application Encoder - Civil Service Form No. 6 (Revised 2020)
Complete version with proper PDF layout matching the uploaded document
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from tkcalendar import DateEntry
from datetime import datetime, date
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib.units import mm
from reportlab.lib import colors
from reportlab.platypus import Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import os

class LeaveApplicationEncoder:
    def __init__(self, root):
        self.root = root
        self.root.title("Leave Application Encoder - Civil Service Form No. 6")
        self.root.geometry("700x650")
        
        # Position options with roman numerals preserved
        self.positions = [
            "Administrative Aide I",
            "Administrative Aide II",
            "Administrative Aide III",
            "Administrative Aide IV",
            "Administrative Aide V",
            "Administrative Aide VI",
            "Social Insurance Assistant I",
            "Social Insurance Assistant II",
            "Social Insurance Assistant III",
            "Social Insurance Assistant IV",
            "Social Insurance Assistant V",
            "Social Insurance Officer I",
            "Social Insurance Officer II",
            "Social Insurance Officer III",
            "Chief Social Insurance Officer I"
        ]
        
        self.create_widgets()
        
    def create_widgets(self):
        # Main frame with scrollbar
        main_canvas = tk.Canvas(self.root)
        scrollbar = ttk.Scrollbar(self.root, orient="vertical", command=main_canvas.yview)
        scrollable_frame = ttk.Frame(main_canvas)
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: main_canvas.configure(scrollregion=main_canvas.bbox("all"))
        )
        
        main_canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        main_canvas.configure(yscrollcommand=scrollbar.set)
        
        main_frame = ttk.Frame(scrollable_frame, padding="20")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Title section - preserving exact Excel text
        title_frame = ttk.LabelFrame(main_frame, text="Form Header", padding="10")
        title_frame.grid(row=0, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(0, 15))
        
        ttk.Label(title_frame, text="CIVIL SERVICE FORM NO. 6", 
                 font=("Arial", 12, "bold")).grid(row=0, column=0, pady=2)
        ttk.Label(title_frame, text="(Revised 2020)").grid(row=1, column=0)
        ttk.Label(title_frame, text="APPLICATION FOR LEAVE", 
                 font=("Arial", 11, "bold")).grid(row=2, column=0, pady=2)
        
        # Office/Department - preserving default value "LHIO - Digos"
        ttk.Label(main_frame, text="Office/Department:").grid(row=1, column=0, sticky=tk.W, pady=5)
        self.office_var = tk.StringVar(value="LHIO - Digos")
        ttk.Entry(main_frame, textvariable=self.office_var, width=40).grid(
            row=1, column=1, sticky=(tk.W, tk.E), pady=5)
        
        # Name field
        ttk.Label(main_frame, text="Name:").grid(row=2, column=0, sticky=tk.W, pady=5)
        self.name_var = tk.StringVar()
        ttk.Entry(main_frame, textvariable=self.name_var, width=40).grid(
            row=2, column=1, sticky=(tk.W, tk.E), pady=5)
        
        # Date of Filing - auto-populated with today's date
        ttk.Label(main_frame, text="Date of Filing:").grid(row=3, column=0, sticky=tk.W, pady=5)
        self.date_filing = DateEntry(main_frame, width=37, background='darkblue',
                                     foreground='white', borderwidth=2, 
                                     date_pattern='mm/dd/yyyy')
        self.date_filing.grid(row=3, column=1, sticky=(tk.W, tk.E), pady=5)
        
        # Position dropdown with exact options
        ttk.Label(main_frame, text="Position:").grid(row=4, column=0, sticky=tk.W, pady=5)
        self.position_var = tk.StringVar()
        position_combo = ttk.Combobox(main_frame, textvariable=self.position_var,
                                     values=self.positions, state="readonly", width=37)
        position_combo.grid(row=4, column=1, sticky=(tk.W, tk.E), pady=5)
        
        # Salary field
        ttk.Label(main_frame, text="Salary:").grid(row=5, column=0, sticky=tk.W, pady=5)
        self.salary_var = tk.StringVar()
        ttk.Entry(main_frame, textvariable=self.salary_var, width=40).grid(
            row=5, column=1, sticky=(tk.W, tk.E), pady=5)
        
        # Inclusive Dates section
        dates_frame = ttk.LabelFrame(main_frame, text="Inclusive Dates", padding="10")
        dates_frame.grid(row=6, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=10)
        
        ttk.Label(dates_frame, text="Start Date:").grid(row=0, column=0, sticky=tk.W, pady=5)
        self.start_date = DateEntry(dates_frame, width=18, background='darkblue',
                                    foreground='white', borderwidth=2, 
                                    date_pattern='mm/dd/yyyy')
        self.start_date.grid(row=0, column=1, sticky=(tk.W, tk.E), pady=5, padx=(5, 0))
        
        ttk.Label(dates_frame, text="End Date:").grid(row=1, column=0, sticky=tk.W, pady=5)
        self.end_date = DateEntry(dates_frame, width=18, background='darkblue',
                                  foreground='white', borderwidth=2,
                                  date_pattern='mm/dd/yyyy')
        self.end_date.grid(row=1, column=1, sticky=(tk.W, tk.E), pady=5, padx=(5, 0))
        
        # Working Days dropdown (1-31 with "day/days" format)
        ttk.Label(main_frame, text="Working Days:").grid(row=7, column=0, sticky=tk.W, pady=5)
        working_days_options = [f"{i} day" if i == 1 else f"{i} days" for i in range(1, 32)]
        self.working_days_var = tk.StringVar()
        working_days_combo = ttk.Combobox(main_frame, textvariable=self.working_days_var,
                                         values=working_days_options, state="readonly", width=37)
        working_days_combo.grid(row=7, column=1, sticky=(tk.W, tk.E), pady=5)
        
        # Buttons frame
        button_frame = ttk.Frame(main_frame)
        button_frame.grid(row=8, column=0, columnspan=2, pady=20)
        
        ttk.Button(button_frame, text="Generate PDF", 
                  command=self.generate_pdf).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Save to Excel", 
                  command=self.save_to_excel).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Clear Form", 
                  command=self.clear_form).pack(side=tk.LEFT, padx=5)
        
        # Configure grid weights
        main_frame.columnconfigure(1, weight=1)
        dates_frame.columnconfigure(1, weight=1)
        
        # Pack canvas and scrollbar
        main_canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
    
    def validate_inputs(self):
        """Validate required fields"""
        if not self.name_var.get().strip():
            messagebox.showerror("Validation Error", "Name is required")
            return False
        if not self.position_var.get():
            messagebox.showerror("Validation Error", "Position must be selected")
            return False
        if not self.working_days_var.get():
            messagebox.showerror("Validation Error", "Working Days must be selected")
            return False
        return True
    
    def save_to_excel(self):
        """Save data to Excel file preserving exact structure"""
        if not self.validate_inputs():
            return
        
        try:
            # Load the template
            wb = openpyxl.load_workbook('ALA.xlsx')
            ws = wb.active
            
            # Populate cells - preserving exact cell locations
            ws['B7'] = self.office_var.get()
            ws['B8'] = self.name_var.get()
            ws['G8'] = self.date_filing.get_date().strftime("%m/%d/%Y")
            ws['B9'] = self.position_var.get()
            ws['G9'] = self.salary_var.get()
            
            # Format inclusive dates
            start = self.start_date.get_date()
            end = self.end_date.get_date()
            if start == end:
                date_range = start.strftime("%B %d, %Y")
            else:
                date_range = f"{start.strftime('%B %d, %Y')} - {end.strftime('%B %d, %Y')}"
            ws['B12'] = date_range
            ws['G12'] = self.working_days_var.get()
            
            # Save with timestamp
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"Leave_Application_{timestamp}.xlsx"
            wb.save(filename)
            wb.close()
            
            messagebox.showinfo("Success", f"Saved to {filename}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save Excel: {str(e)}")
    
    def generate_pdf(self):
        """Generate PDF with complete layout matching the uploaded document structure"""
        if not self.validate_inputs():
            return
        
        try:
            # Create output directory
            os.makedirs("output_pdf", exist_ok=True)
            
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"output_pdf/Application_for_Leave_{timestamp}.pdf"
            
            # Create PDF
            c = canvas.Canvas(filename, pagesize=A4)
            width, height = A4
            
            # Draw header section (matching uploaded document structure)
            self.draw_pdf_header(c, width, height)
            
            # Draw form content
            self.draw_pdf_form(c, width, height)
            
            # Save PDF
            c.save()
            
            messagebox.showinfo("Success", f"PDF generated: {filename}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to generate PDF: {str(e)}")
    
    def draw_pdf_header(self, c, width, height):
        """Draw the PDF header section with proper formatting"""
        y_pos = height - 30*mm
        
        # Header text - matching uploaded document
        c.setFont("Helvetica", 9)
        c.drawCentredString(width/2, y_pos, "Republic of the Philippines")
        y_pos -= 4*mm
        
        c.setFont("Helvetica-Bold", 10)
        c.drawCentredString(width/2, y_pos, "PHILIPPINE HEALTH INSURANCE CORPORATION")
        y_pos -= 4*mm
        
        c.setFont("Helvetica-Bold", 9)
        c.drawCentredString(width/2, y_pos, "PhilHealth Regional Office XI")
        y_pos -= 4*mm
        
        c.setFont("Helvetica", 8)
        c.drawCentredString(width/2, y_pos, "J.P. Laurel Avenue, Bajada, Poblacion District, Davao City")
        y_pos -= 3*mm
        c.drawCentredString(width/2, y_pos, "(082) 295-2133 local 6000, (082) 295-3385")
        y_pos -= 3*mm
        c.drawCentredString(width/2, y_pos, "✉ teamphilhealth11 ✉ teamphilhealth @ www.philhealth.gov.ph")
        y_pos -= 8*mm
        
        # Stamp of Date of Receipt box (top right)
        stamp_x = width - 50*mm
        stamp_y = height - 25*mm
        c.rect(stamp_x, stamp_y, 40*mm, 15*mm)
        c.setFont("Helvetica", 7)
        c.drawString(stamp_x + 2*mm, stamp_y + 12*mm, "Stamp of Date of Receipt")
        
        # Form title section
        c.setFont("Helvetica-Bold", 10)
        c.drawCentredString(width/2, y_pos, "Civil Service Form No. 6")
        y_pos -= 4*mm
        c.setFont("Helvetica", 9)
        c.drawCentredString(width/2, y_pos, "Revised 2020")
        y_pos -= 6*mm
        
        c.setFont("Helvetica-Bold", 11)
        c.drawCentredString(width/2, y_pos, "APPLICATION FOR LEAVE")
        
        return y_pos - 10*mm  # Return position for form content
    
    def draw_pdf_form(self, c, width, height):
        """Draw the form fields section"""
        # Starting position after header
        y_pos = height - 100*mm
        left_margin = 20*mm
        label_width = 40*mm
        
        c.setFont("Helvetica", 10)
        
        # 1. OFFICE/DEPARTMENT
        c.drawString(left_margin, y_pos, "1.   OFFICE/DEPARTMENT")
        c.drawString(left_margin + 60*mm, y_pos, "2.  NAME :            (Last)                               (First)                         (Middle)")
        y_pos -= 6*mm
        
        # Office value
        c.setFont("Helvetica-Bold", 10)
        c.drawString(left_margin + 10*mm, y_pos, self.office_var.get())
        y_pos -= 8*mm
        
        # 3. DATE OF FILING and 4. POSITION and 5. SALARY
        c.setFont("Helvetica", 10)
        c.drawString(left_margin, y_pos, f"3.   DATE OF FILING  {self.date_filing.get_date().strftime('%B %d, %Y')}")
        c.drawString(left_margin + 80*mm, y_pos, f"4.   POSITION  {self.position_var.get()}")
        c.drawString(left_margin + 140*mm, y_pos, f"5.  SALARY  {self.salary_var.get()}")
        y_pos -= 10*mm
        
        # 6. DETAILS OF APPLICATION header
        c.setFont("Helvetica-Bold", 10)
        c.drawString(left_margin, y_pos, "6.  DETAILS OF APPLICATION")
        y_pos -= 6*mm
        
        # 6.A TYPE OF LEAVE and 6.B DETAILS OF LEAVE (side by side)
        c.setFont("Helvetica", 9)
        c.drawString(left_margin, y_pos, "6.A  TYPE OF LEAVE TO BE AVAILED OF")
        c.drawString(left_margin + 90*mm, y_pos, "6.B  DETAILS OF LEAVE")
        y_pos -= 5*mm
        
        # Leave types (left column)
        leave_types = [
            "Vacation Leave (Sec. 51, Rule XVI, Omnibus Rules Implementing E.O. No. 292)",
            "Mandatory/Forced Leave(Sec. 25, Rule XVI, Omnibus Rules Implementing E.O. No. 292)",
            "Sick Leave  (Sec. 43, Rule XVI, Omnibus Rules Implementing E.O. No. 292)",
            "Maternity Leave (R.A. No. 11210 / IRR issued by CSC, DOLE and SSS)",
            "Paternity Leave (R.A. No. 8187 / CSC MC No. 71, s. 1998, as amended)",
            "Special Privilege Leave (Sec. 21, Rule XVI, Omnibus Rules Implementing E.O. No. 292)",
            "Solo Parent Leave (RA No. 8972 / CSC MC No. 8, s. 2004)",
            "Study Leave (Sec. 68, Rule XVI, Omnibus Rules Implementing E.O. No. 292)",
            "10-Day VAWC Leave (RA No. 9262 / CSC MC No. 15, s. 2005)",
            "Rehabilitation Privilege (Sec. 55, Rule XVI, Omnibus Rules Implementing E.O. No. 292)",
            "Special Leave Benefits for Women (RA No. 9710 / CSC MC No. 25, s. 2010)",
            "Special Emergency (Calamity) Leave (CSC MC No. 2, s. 2012, as amended)",
            "Adoption Leave (R.A. No. 8552)"
        ]
        
        c.setFont("Helvetica", 7)
        for leave_type in leave_types[:8]:  # First 8 types
            c.rect(left_margin + 5*mm, y_pos - 1*mm, 3*mm, 3*mm)  # Checkbox
            c.drawString(left_margin + 10*mm, y_pos, leave_type[:60])  # Truncate if too long
            y_pos -= 4*mm
        
        # 6.C NUMBER OF WORKING DAYS APPLIED FOR
        y_pos -= 5*mm
        c.setFont("Helvetica", 9)
        c.drawString(left_margin, y_pos, "6.C  NUMBER OF WORKING DAYS APPLIED FOR")
        c.drawString(left_margin + 90*mm, y_pos, "6.D  COMMUTATION")
        y_pos -= 5*mm
        
        # Working days value
        c.setFont("Helvetica-Bold", 10)
        c.drawString(left_margin + 10*mm, y_pos, self.working_days_var.get())
        
        # Commutation checkboxes
        c.setFont("Helvetica", 9)
        c.rect(left_margin + 90*mm, y_pos - 1*mm, 3*mm, 3*mm)
        c.drawString(left_margin + 95*mm, y_pos, "Not Requested")
        y_pos -= 4*mm
        c.rect(left_margin + 90*mm, y_pos - 1*mm, 3*mm, 3*mm)
        c.drawString(left_margin + 95*mm, y_pos, "Requested")
        y_pos -= 6*mm
        
        # INCLUSIVE DATES
        c.setFont("Helvetica", 9)
        c.drawString(left_margin + 10*mm, y_pos, "INCLUSIVE DATES")
        y_pos -= 4*mm
        
        # Format dates
        start = self.start_date.get_date()
        end = self.end_date.get_date()
        if start == end:
            date_range = start.strftime("%B %d, %Y")
        else:
            date_range = f"{start.strftime('%B %d, %Y')} - {end.strftime('%B %d, %Y')}"
        
        c.setFont("Helvetica-Bold", 10)
        c.drawString(left_margin + 10*mm, y_pos, date_range)
        y_pos -= 8*mm
        
        # Signature line
        c.setFont("Helvetica", 8)
        c.line(left_margin + 90*mm, y_pos, left_margin + 140*mm, y_pos)
        y_pos -= 3*mm
        c.drawCentredString(left_margin + 115*mm, y_pos, "(Signature of Applicant)")
        y_pos -= 10*mm
        
        # 7. DETAILS OF ACTION ON APPLICATION
        c.setFont("Helvetica-Bold", 10)
        c.drawString(left_margin, y_pos, "7.  DETAILS OF ACTION ON APPLICATION")
        y_pos -= 6*mm
        
        # 7.A CERTIFICATION OF LEAVE CREDITS and 7.B RECOMMENDATION
        c.setFont("Helvetica", 9)
        c.drawString(left_margin, y_pos, "7.A  CERTIFICATION OF LEAVE CREDITS")
        c.drawString(left_margin + 90*mm, y_pos, "7.B  RECOMMENDATION")
        y_pos -= 5*mm
        
        # Leave credits table
        c.drawString(left_margin + 5*mm, y_pos, "As of _______________________")
        c.drawString(left_margin + 90*mm, y_pos, "For approval")
        y_pos -= 5*mm
        
        # Table headers
        c.drawString(left_margin + 20*mm, y_pos, "Vacation Leave")
        c.drawString(left_margin + 45*mm, y_pos, "Sick Leave")
        y_pos -= 4*mm
        c.drawString(left_margin + 5*mm, y_pos, "Total Earned")
        c.drawString(left_margin + 90*mm, y_pos, "For disapproval due to ________________________")
        y_pos -= 4*mm
        c.drawString(left_margin + 5*mm, y_pos, "Less this application")
        y_pos -= 4*mm
        c.drawString(left_margin + 5*mm, y_pos, "Balance")
        y_pos -= 8*mm
        
        # Authorized Officer signatures
        c.line(left_margin + 5*mm, y_pos, left_margin + 50*mm, y_pos)
        c.line(left_margin + 90*mm, y_pos, left_margin + 135*mm, y_pos)
        y_pos -= 3*mm
        c.setFont("Helvetica", 8)
        c.drawString(left_margin + 15*mm, y_pos, "(Authorized Officer)")
        c.drawString(left_margin + 100*mm, y_pos, "(Authorized Officer)")
        y_pos -= 8*mm
        
        # 7.C APPROVED FOR and 7.D DISAPPROVED DUE TO
        c.setFont("Helvetica", 9)
        c.drawString(left_margin, y_pos, "7.C  APPROVED FOR:")
        c.drawString(left_margin + 90*mm, y_pos, "7.D   DISAPPROVED DUE TO:")
        y_pos -= 4*mm
        
        c.drawString(left_margin + 10*mm, y_pos, "_______ days with pay")
        y_pos -= 4*mm
        c.drawString(left_margin + 10*mm, y_pos, "_______ days without pay")
        y_pos -= 4*mm
        c.drawString(left_margin + 10*mm, y_pos, "_______ others (Specify)")
        y_pos -= 10*mm
        
        # Final signature line
        c.line(left_margin + 60*mm, y_pos, left_margin + 110*mm, y_pos)
        y_pos -= 3*mm
        c.setFont("Helvetica", 8)
        c.drawCentredString(left_margin + 85*mm, y_pos, "(Authorized Official)")
    
    def clear_form(self):
        """Clear all form fields except Office/Department default"""
        self.name_var.set("")
        self.position_var.set("")
        self.salary_var.set("")
        self.working_days_var.set("")
        self.date_filing.set_date(datetime.now())
        self.start_date.set_date(datetime.now())
        self.end_date.set_date(datetime.now())
        # Keep Office/Department as "LHIO - Digos"

def main():
    root = tk.Tk()
    app = LeaveApplicationEncoder(root)
    root.mainloop()

if __name__ == "__main__":
    main()
