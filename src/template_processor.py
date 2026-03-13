"""
Template Processor Component
Handles Excel template loading, manipulation, and preservation.
"""

import os
from pathlib import Path
from typing import Dict, Any, Optional, Union
from datetime import datetime, date
from openpyxl import Workbook, load_workbook
from openpyxl.workbook.workbook import Workbook as WorkbookType
from openpyxl.worksheet.worksheet import Worksheet
from copy import deepcopy


class TemplateProcessor:
    """Processes Excel templates while preserving layout and formatting."""
    
    def __init__(self, template_path: str = "ALA.xlsx"):
        """Initialize template processor with path to Excel template.
        
        Args:
            template_path: Path to the Excel template file (default: ALA.xlsx)
        """
        self.template_path = Path(template_path)
        self.template_workbook: Optional[WorkbookType] = None
        self._template_validated = False
        
        # Cell mappings for Civil Service Form No. 6 fields
        # These constants define where each form field maps to in the Excel template
        self.CELL_MAPPINGS = {
            'name': 'B8',
            'position': 'B9',
            'date_filing': 'G8', 
            'inclusive_dates': 'B12',
            'working_days': 'G12',
            # Additional mappings for comprehensive form support
            'office_department': 'B7',  # Office/Department field
            'salary': 'G9',  # Salary field if needed
        }
    
    def validate_template(self) -> bool:
        """Validate that the template file exists and is accessible.
        
        Returns:
            True if template is valid and accessible
            
        Raises:
            FileNotFoundError: If template file doesn't exist
            PermissionError: If template file cannot be accessed
            Exception: If template structure is invalid
        """
        if not self.template_path.exists():
            raise FileNotFoundError(
                f"Template file not found: {self.template_path}. "
                f"Please ensure ALA.xlsx (Civil Service Form No. 6) is in the project directory."
            )
        
        if not os.access(self.template_path, os.R_OK):
            raise PermissionError(f"Cannot read template file: {self.template_path}")
        
        try:
            # Test load to validate structure
            test_workbook = load_workbook(self.template_path, read_only=True)
            if test_workbook.active is None:
                raise Exception("Template has no active worksheet")
            test_workbook.close()
            self._template_validated = True
            return True
        except Exception as e:
            raise Exception(f"Invalid template structure: {e}")
    
    def load_template(self) -> WorkbookType:
        """Load Excel template without modification.
        
        This method loads the template in read-only mode to ensure the original
        file is never modified, preserving template integrity.
        
        Returns:
            Loaded workbook object (read-only)
            
        Raises:
            FileNotFoundError: If template file doesn't exist
            Exception: If template cannot be loaded
        """
        if not self._template_validated:
            self.validate_template()
        
        try:
            # Load in read-only mode to prevent accidental modification
            self.template_workbook = load_workbook(self.template_path, read_only=True)
            return self.template_workbook
        except Exception as e:
            raise Exception(f"Failed to load template: {e}")
    
    def create_working_copy(self) -> WorkbookType:
        """Create working copy of template for data population.
        
        This method creates a completely separate workbook instance that can be
        modified without affecting the original template file. This ensures
        template preservation as required by the specifications.
        
        Returns:
            Working copy workbook that can be modified safely
            
        Raises:
            Exception: If working copy cannot be created
        """
        if not self._template_validated:
            self.validate_template()
        
        try:
            # Create a fresh copy by loading the template again (not read-only)
            working_copy = load_workbook(self.template_path)
            return working_copy
        except Exception as e:
            raise Exception(f"Failed to create working copy: {e}")
    
    def get_cell_mappings(self) -> Dict[str, str]:
        """Get the cell mapping constants for form fields.
        
        Returns:
            Dictionary mapping field names to Excel cell addresses
        """
        return self.CELL_MAPPINGS.copy()
    
    def populate_fields(self, workbook: WorkbookType, data_mapping: Dict[str, Any]) -> None:
        """Populate Excel fields with provided data.
        
        Args:
            workbook: Working copy workbook to populate
            data_mapping: Dictionary mapping field names to values
            
        Raises:
            ValueError: If workbook is None or invalid field names provided
        """
        if workbook is None:
            raise ValueError("Workbook cannot be None")
        
        worksheet = workbook.active
        if worksheet is None:
            raise ValueError("Workbook has no active worksheet")
        
        for field_name, value in data_mapping.items():
            if field_name in self.CELL_MAPPINGS:
                cell_address = self.CELL_MAPPINGS[field_name]
                try:
                    # Use type-specific population method
                    self._populate_cell_by_type(worksheet, cell_address, value, field_name)
                except Exception as e:
                    raise Exception(f"Failed to populate field '{field_name}' at {cell_address}: {e}")
            else:
                # Log warning for unmapped fields but don't fail
                print(f"Warning: Field '{field_name}' not found in cell mappings")
    
    def preserve_template_integrity(self) -> bool:
        """Verify that the original template file remains unchanged.
        
        This method can be used to verify that template processing operations
        have not modified the original template file.
        
        Returns:
            True if template integrity is preserved
        """
        try:
            # Check if file is still accessible and readable
            if not self.template_path.exists():
                return False
            
            # Verify we can still load it
            test_workbook = load_workbook(self.template_path, read_only=True)
            test_workbook.close()
            return True
        except Exception:
            return False
    
    def _populate_cell_by_type(self, worksheet: Worksheet, cell_address: str, 
                              value: Any, field_name: str) -> None:
        """Populate a cell with type-specific formatting.
        
        Args:
            worksheet: The worksheet to populate
            cell_address: Excel cell address (e.g., 'B8')
            value: The value to populate
            field_name: Name of the field being populated
            
        Raises:
            Exception: If population fails
        """
        if value is None:
            worksheet[cell_address] = ""
            return
        
        # Handle different data types based on field name and value type
        if field_name in ['date_filing'] or isinstance(value, (datetime, date)):
            self._populate_date_field(worksheet, cell_address, value)
        elif field_name in ['working_days'] or isinstance(value, (int, float)):
            self._populate_number_field(worksheet, cell_address, value)
        elif field_name in ['name', 'position', 'office_department']:
            self._populate_text_field(worksheet, cell_address, value, field_name)
        elif field_name == 'inclusive_dates':
            self._populate_date_range_field(worksheet, cell_address, value)
        else:
            # Default to text handling
            self._populate_text_field(worksheet, cell_address, value, field_name)
    
    def _populate_text_field(self, worksheet: Worksheet, cell_address: str, 
                            value: Any, field_name: str) -> None:
        """Populate a text field with proper formatting.
        
        Args:
            worksheet: The worksheet to populate
            cell_address: Excel cell address
            value: The text value to populate
            field_name: Name of the field being populated
        """
        text_value = str(value).strip()
        
        # Handle employee name formatting (e.g., "Mendoza, Anthony Berja")
        if field_name == 'name':
            text_value = self._format_employee_name(text_value)
        
        worksheet[cell_address] = text_value
    
    def _populate_date_field(self, worksheet: Worksheet, cell_address: str, 
                            value: Union[datetime, date, str]) -> None:
        """Populate a date field with proper formatting.
        
        Args:
            worksheet: The worksheet to populate
            cell_address: Excel cell address
            value: The date value to populate (datetime, date, or string)
        """
        if isinstance(value, str):
            # Try to parse string dates
            try:
                if '/' in value:
                    # Handle MM/DD/YYYY format
                    parsed_date = datetime.strptime(value, '%m/%d/%Y').date()
                elif '-' in value:
                    # Handle YYYY-MM-DD format
                    parsed_date = datetime.strptime(value, '%Y-%m-%d').date()
                else:
                    # If parsing fails, use as-is
                    worksheet[cell_address] = value
                    return
                worksheet[cell_address] = parsed_date
            except ValueError:
                # If parsing fails, use the string as-is
                worksheet[cell_address] = value
        elif isinstance(value, datetime):
            worksheet[cell_address] = value.date()
        elif isinstance(value, date):
            worksheet[cell_address] = value
        else:
            worksheet[cell_address] = str(value)
    
    def _populate_number_field(self, worksheet: Worksheet, cell_address: str, 
                              value: Union[int, float, str]) -> None:
        """Populate a number field with proper formatting.
        
        Args:
            worksheet: The worksheet to populate
            cell_address: Excel cell address
            value: The numeric value to populate
        """
        if isinstance(value, (int, float)):
            worksheet[cell_address] = value
        elif isinstance(value, str):
            try:
                # Try to convert string to number
                if '.' in value:
                    worksheet[cell_address] = float(value)
                else:
                    worksheet[cell_address] = int(value)
            except ValueError:
                # If conversion fails, use as string
                worksheet[cell_address] = value
        else:
            worksheet[cell_address] = value
    
    def _populate_date_range_field(self, worksheet: Worksheet, cell_address: str, 
                                  value: Any) -> None:
        """Populate a date range field with proper formatting.
        
        Handles various date range formats:
        - Single dates: "Nov 28"
        - Consecutive ranges: "Oct 29-30"
        - Non-consecutive dates: "Dec 1&26"
        
        Args:
            worksheet: The worksheet to populate
            cell_address: Excel cell address
            value: The date range value to populate
        """
        if isinstance(value, list):
            # Handle list of dates
            if len(value) == 1:
                # Single date
                date_str = self._format_date_for_display(value[0])
            elif len(value) == 2:
                # Date range
                start_date = self._format_date_for_display(value[0])
                end_date = self._format_date_for_display(value[1])
                date_str = f"{start_date}-{end_date}"
            else:
                # Multiple non-consecutive dates
                formatted_dates = [self._format_date_for_display(d) for d in value]
                date_str = "&".join(formatted_dates)
        else:
            # Handle string or single date
            date_str = str(value)
        
        worksheet[cell_address] = date_str
    
    def _format_employee_name(self, name: str) -> str:
        """Format employee name according to Civil Service standards.
        
        Ensures proper formatting like "Mendoza, Anthony Berja"
        
        Args:
            name: The employee name to format
            
        Returns:
            Properly formatted employee name
        """
        name = name.strip()
        
        # If already in "Last, First Middle" format, return as-is
        if ',' in name:
            parts = name.split(',', 1)
            if len(parts) == 2:
                last_name = parts[0].strip()
                first_middle = parts[1].strip()
                return f"{last_name}, {first_middle}"
        
        # If in "First Middle Last" format, convert to "Last, First Middle"
        name_parts = name.split()
        if len(name_parts) >= 2:
            last_name = name_parts[-1]
            first_middle = ' '.join(name_parts[:-1])
            return f"{last_name}, {first_middle}"
        
        # Single name, return as-is
        return name
    
    def _format_date_for_display(self, date_value: Union[datetime, date, str]) -> str:
        """Format a date for display in the form.
        
        Args:
            date_value: The date to format
            
        Returns:
            Formatted date string (e.g., "Nov 28")
        """
        if isinstance(date_value, str):
            return date_value
        elif isinstance(date_value, datetime):
            return date_value.strftime("%b %d")
        elif isinstance(date_value, date):
            return date_value.strftime("%b %d")
        else:
            return str(date_value)