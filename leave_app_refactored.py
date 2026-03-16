#!/usr/bin/env python3
"""
Leave Application Encoder - Refactored Version
Civil Service Form No. 6 (Revised 2020) - PhilHealth Regional Office XI Standard
Grid-based layout with precise positioning and professional formatting
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from tkcalendar import DateEntry
from datetime import datetime, date
import openpyxl
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib.units import mm, inch
from reportlab.lib import colors
from reportlab.lib.utils import ImageReader
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import os
import textwrap
from PIL import Image

# GLOBAL GRID SYSTEM CONSTANTS
CANVAS_WIDTH = A4[0]  # 595.276 points
CANVAS_HEIGHT = A4[1]  # 841.890 points
MARGIN_LEFT = 20*mm
MARGIN_RIGHT = 20*mm
MARGIN_TOP = 20*mm
MARGIN_BOTTOM = 20*mm

# SECTION Y-COORDINATES (hardcoded to prevent overlap)
HEADER_HEIGHT = 150  # pixels for logos
SECTION_1_Y = CANVAS_HEIGHT - 100*mm  # Office/Department & Name (sections 1-5) - FIXED: Reduced gap
SECTION_3_Y = CANVAS_HEIGHT - 140*mm  # Details of Application header
SECTION_6A_Y = CANVAS_HEIGHT - 160*mm  # Type of Leave (Column A)
SECTION_6B_Y = CANVAS_HEIGHT - 160*mm  # Details of Leave (Column B)
SECTION_6C_Y = CANVAS_HEIGHT - 320*mm  # Working Days & Commutation
SECTION_7_Y = CANVAS_HEIGHT - 420*mm   # Details of Action

# COLUMN DEFINITIONS for Section 6
COLUMN_A_START = MARGIN_LEFT
COLUMN_A_END = CANVAS_WIDTH * 0.55
COLUMN_B_START = CANVAS_WIDTH * 0.56
COLUMN_B_END = CANVAS_WIDTH - MARGIN_RIGHT

# FONT SIZES
FONT_HEADER = 12
FONT_SECTION = 10
FONT_LABEL = 9
FONT_DATA = 10
FONT_SMALL = 8

class LeaveApplicationEncoder:
    def __init__(self, root):
        self.root = root
        self.root.title("Leave Application Encoder - PhilHealth Regional Office XI")
        self.root.geometry("800x700")
        
        # Data model structure - nested dictionary
        self.form_data = {
            'office_department': 'LHIO - Digos',
            'name': {'last': '', 'first': '', 'middle': ''},
            'date_filing': datetime.now(),
            'position': '',
            'salary': '',
            'leave_types': {
                'vacation': False,
                'mandatory': False,
                'sick': False,
                'maternity': False,
                'paternity': False,
                'special_privilege': False,
                'solo_parent': False,
                'study': False,
                'vawc_10day': False,
                'rehabilitation': False,
                'special_women': False,
                'special_emergency': False,
                'adoption': False
            },
            'inclusive_dates': {'start': datetime.now(), 'end': datetime.now()},
            'working_days': 1,
            'commutation': {'not_requested': True, 'requested': False}
        }
        
        # Position options with roman numerals preserved
        self.positions = [
            "Administrative Aide I", "Administrative Aide II", "Administrative Aide III",
            "Administrative Aide IV", "Administrative Aide V", "Administrative Aide VI",
            "Social Insurance Assistant I", "Social Insurance Assistant II", 
            "Social Insurance Assistant III", "Social Insurance Assistant IV", 
            "Social Insurance Assistant V", "Social Insurance Officer I", 
            "Social Insurance Officer II", "Social Insurance Officer III",
            "Chief Social Insurance Officer I"
        ]
        
        self.create_widgets()
        
    def create_widgets(self):
        # Main scrollable frame
        main_canvas = tk.Canvas(self.root)
        scrollbar = ttk.Scrollbar(self.root, orient="vertical", command=main_canvas.yview)
        scrollable_frame = ttk.Frame(main_canvas)
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: main_canvas.configure(scrollregion=main_canvas.bbox("all"))
        )
        
        main_canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        main_canvas.configure(yscrollcommand=scrollbar.set)
        
        main_frame = ttk.Frame(scrollable_frame, padding="20")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Header section with logo configuration
        header_frame = ttk.LabelFrame(main_frame, text="Header Configuration", padding="10")
        header_frame.grid(row=0, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(0, 15))
        
        ttk.Label(header_frame, text="PhilHealth Logo:").grid(row=0, column=0, sticky=tk.W, pady=2)
        self.philhealth_logo_var = tk.StringVar(value="logos/philhealth_logo.png")
        ttk.Entry(header_frame, textvariable=self.philhealth_logo_var, width=30).grid(row=0, column=1, padx=5)
        ttk.Button(header_frame, text="Browse", command=lambda: self.browse_logo('philhealth')).grid(row=0, column=2)
        
        ttk.Label(header_frame, text="Bagong Pilipinas Logo:").grid(row=1, column=0, sticky=tk.W, pady=2)
        self.bagong_logo_var = tk.StringVar(value="logos/bagong_pilipinas_logo.png")
        ttk.Entry(header_frame, textvariable=self.bagong_logo_var, width=30).grid(row=1, column=1, padx=5)
        ttk.Button(header_frame, text="Browse", command=lambda: self.browse_logo('bagong')).grid(row=1, column=2)
        
        # Form data section
        form_frame = ttk.LabelFrame(main_frame, text="Form Data", padding="10")
        form_frame.grid(row=1, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(0, 15))
        
        # Office/Department
        ttk.Label(form_frame, text="1. OFFICE/DEPARTMENT:").grid(row=0, column=0, sticky=tk.W, pady=5)
        self.office_var = tk.StringVar(value=self.form_data['office_department'])
        ttk.Entry(form_frame, textvariable=self.office_var, width=40).grid(row=0, column=1, columnspan=2, sticky=(tk.W, tk.E), pady=5)
        
        # Name fields (Last, First, Middle)
        ttk.Label(form_frame, text="2. NAME:").grid(row=1, column=0, sticky=tk.W, pady=5)
        name_frame = ttk.Frame(form_frame)
        name_frame.grid(row=1, column=1, columnspan=2, sticky=(tk.W, tk.E), pady=5)
        
        ttk.Label(name_frame, text="Last:").grid(row=0, column=0, sticky=tk.W)
        self.name_last_var = tk.StringVar()
        ttk.Entry(name_frame, textvariable=self.name_last_var, width=15).grid(row=0, column=1, padx=2)
        
        ttk.Label(name_frame, text="First:").grid(row=0, column=2, sticky=tk.W, padx=(10,0))
        self.name_first_var = tk.StringVar()
        ttk.Entry(name_frame, textvariable=self.name_first_var, width=15).grid(row=0, column=3, padx=2)
        
        ttk.Label(name_frame, text="Middle:").grid(row=0, column=4, sticky=tk.W, padx=(10,0))
        self.name_middle_var = tk.StringVar()
        ttk.Entry(name_frame, textvariable=self.name_middle_var, width=15).grid(row=0, column=5, padx=2)
        
        # Date of Filing
        ttk.Label(form_frame, text="3. DATE OF FILING:").grid(row=2, column=0, sticky=tk.W, pady=5)
        self.date_filing = DateEntry(form_frame, width=20, background='darkblue',
                                     foreground='white', borderwidth=2, date_pattern='mm/dd/yyyy')
        self.date_filing.grid(row=2, column=1, sticky=tk.W, pady=5)
        
        # Position with character limit
        ttk.Label(form_frame, text="4. POSITION:").grid(row=3, column=0, sticky=tk.W, pady=5)
        self.position_var = tk.StringVar()
        position_combo = ttk.Combobox(form_frame, textvariable=self.position_var,
                                     values=self.positions, state="readonly", width=35)
        position_combo.grid(row=3, column=1, columnspan=2, sticky=(tk.W, tk.E), pady=5)
        
        # Salary with character limit
        ttk.Label(form_frame, text="5. SALARY:").grid(row=4, column=0, sticky=tk.W, pady=5)
        self.salary_var = tk.StringVar()
        salary_entry = ttk.Entry(form_frame, textvariable=self.salary_var, width=20)
        salary_entry.grid(row=4, column=1, sticky=tk.W, pady=5)
        
        # Leave types section (6.A)
        leave_frame = ttk.LabelFrame(main_frame, text="6.A TYPE OF LEAVE TO BE AVAILED OF", padding="10")
        leave_frame.grid(row=2, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(0, 15))
        
        # Create checkboxes for all leave types
        self.leave_vars = {}
        leave_types = [
            ('vacation', 'Vacation Leave (Sec. 51, Rule XVI, Omnibus Rules Implementing E.O. No. 292)'),
            ('mandatory', 'Mandatory/Forced Leave (Sec. 25, Rule XVI, Omnibus Rules Implementing E.O. No. 292)'),
            ('sick', 'Sick Leave (Sec. 43, Rule XVI, Omnibus Rules Implementing E.O. No. 292)'),
            ('maternity', 'Maternity Leave (R.A. No. 11210 / IRR issued by CSC, DOLE and SSS)'),
            ('paternity', 'Paternity Leave (R.A. No. 8187 / CSC MC No. 71, s. 1998, as amended)'),
            ('special_privilege', 'Special Privilege Leave (Sec. 21, Rule XVI, Omnibus Rules Implementing E.O. No. 292)'),
            ('solo_parent', 'Solo Parent Leave (RA No. 8972 / CSC MC No. 8, s. 2004)'),
            ('study', 'Study Leave (Sec. 68, Rule XVI, Omnibus Rules Implementing E.O. No. 292)'),
            ('vawc_10day', '10-Day VAWC Leave (RA No. 9262 / CSC MC No. 15, s. 2005)'),
            ('rehabilitation', 'Rehabilitation Privilege (Sec. 55, Rule XVI, Omnibus Rules Implementing E.O. No. 292)'),
            ('special_women', 'Special Leave Benefits for Women (RA No. 9710 / CSC MC No. 25, s. 2010)'),
            ('special_emergency', 'Special Emergency (Calamity) Leave (CSC MC No. 2, s. 2012, as amended)'),
            ('adoption', 'Adoption Leave (R.A. No. 8552)')
        ]
        
        for i, (key, label) in enumerate(leave_types):
            self.leave_vars[key] = tk.BooleanVar()
            cb = ttk.Checkbutton(leave_frame, text=label, variable=self.leave_vars[key])
            cb.grid(row=i, column=0, sticky=tk.W, pady=2)
        
        # Dates and working days section
        dates_frame = ttk.LabelFrame(main_frame, text="6.C & 6.D - Dates and Working Days", padding="10")
        dates_frame.grid(row=3, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(0, 15))
        
        # Inclusive dates
        ttk.Label(dates_frame, text="INCLUSIVE DATES:").grid(row=0, column=0, sticky=tk.W, pady=5)
        dates_sub_frame = ttk.Frame(dates_frame)
        dates_sub_frame.grid(row=0, column=1, columnspan=2, sticky=(tk.W, tk.E), pady=5)
        
        ttk.Label(dates_sub_frame, text="Start:").grid(row=0, column=0, sticky=tk.W)
        self.start_date = DateEntry(dates_sub_frame, width=12, background='darkblue',
                                    foreground='white', borderwidth=2, date_pattern='mm/dd/yyyy')
        self.start_date.grid(row=0, column=1, padx=5)
        
        ttk.Label(dates_sub_frame, text="End:").grid(row=0, column=2, sticky=tk.W, padx=(10,0))
        self.end_date = DateEntry(dates_sub_frame, width=12, background='darkblue',
                                  foreground='white', borderwidth=2, date_pattern='mm/dd/yyyy')
        self.end_date.grid(row=0, column=3, padx=5)
        
        # Working days
        ttk.Label(dates_frame, text="NUMBER OF WORKING DAYS:").grid(row=1, column=0, sticky=tk.W, pady=5)
        working_days_options = [f"{i} day" if i == 1 else f"{i} days" for i in range(1, 32)]
        self.working_days_var = tk.StringVar()
        working_days_combo = ttk.Combobox(dates_frame, textvariable=self.working_days_var,
                                         values=working_days_options, state="readonly", width=15)
        working_days_combo.grid(row=1, column=1, sticky=tk.W, pady=5)
        
        # Commutation
        ttk.Label(dates_frame, text="COMMUTATION:").grid(row=2, column=0, sticky=tk.W, pady=5)
        comm_frame = ttk.Frame(dates_frame)
        comm_frame.grid(row=2, column=1, columnspan=2, sticky=(tk.W, tk.E), pady=5)
        
        self.comm_var = tk.StringVar(value="not_requested")
        ttk.Radiobutton(comm_frame, text="Not Requested", variable=self.comm_var, 
                       value="not_requested").grid(row=0, column=0, sticky=tk.W)
        ttk.Radiobutton(comm_frame, text="Requested", variable=self.comm_var, 
                       value="requested").grid(row=0, column=1, sticky=tk.W, padx=(20,0))
        
        # Buttons
        button_frame = ttk.Frame(main_frame)
        button_frame.grid(row=4, column=0, columnspan=3, pady=20)
        
        ttk.Button(button_frame, text="Generate PDF", command=self.generate_pdf).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Save to Excel", command=self.save_to_excel).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Clear Form", command=self.clear_form).pack(side=tk.LEFT, padx=5)
        
        # Configure grid weights
        main_frame.columnconfigure(1, weight=1)
        form_frame.columnconfigure(1, weight=1)
        leave_frame.columnconfigure(0, weight=1)
        dates_frame.columnconfigure(1, weight=1)
        
        # Pack canvas and scrollbar
        main_canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
    
    def browse_logo(self, logo_type):
        """Browse for logo files"""
        filename = filedialog.askopenfilename(
            title=f"Select {logo_type.title()} Logo",
            filetypes=[("Image files", "*.png *.jpg *.jpeg *.gif *.bmp")]
        )
        if filename:
            if logo_type == 'philhealth':
                self.philhealth_logo_var.set(filename)
            else:
                self.bagong_logo_var.set(filename)
    
    def validate_inputs(self):
        """Validate required fields with enhanced checking"""
        errors = []
        
        if not self.name_last_var.get().strip():
            errors.append("Last name is required")
        if not self.name_first_var.get().strip():
            errors.append("First name is required")
        if not self.position_var.get():
            errors.append("Position must be selected")
        if not self.working_days_var.get():
            errors.append("Working days must be selected")
        
        # Check if at least one leave type is selected
        if not any(var.get() for var in self.leave_vars.values()):
            errors.append("At least one leave type must be selected")
        
        if errors:
            messagebox.showerror("Validation Error", "\n".join(errors))
            return False
        return True
    
    def update_form_data(self):
        """Update the form_data dictionary with current values"""
        self.form_data.update({
            'office_department': self.office_var.get(),
            'name': {
                'last': self.name_last_var.get(),
                'first': self.name_first_var.get(),
                'middle': self.name_middle_var.get()
            },
            'date_filing': self.date_filing.get_date(),
            'position': self.position_var.get(),
            'salary': self.salary_var.get(),
            'leave_types': {key: var.get() for key, var in self.leave_vars.items()},
            'inclusive_dates': {
                'start': self.start_date.get_date(),
                'end': self.end_date.get_date()
            },
            'working_days': int(self.working_days_var.get().split()[0]) if self.working_days_var.get() else 1,
            'commutation': {
                'not_requested': self.comm_var.get() == 'not_requested',
                'requested': self.comm_var.get() == 'requested'
            }
        })
    
    def save_to_excel(self):
        """Save data to Excel file preserving exact structure"""
        if not self.validate_inputs():
            return
        
        self.update_form_data()
        
        try:
            # Load the template
            wb = openpyxl.load_workbook('ALA.xlsx')
            ws = wb.active
            
            # Populate cells with exact mapping
            ws['B7'] = self.form_data['office_department']
            full_name = f"{self.form_data['name']['last']}, {self.form_data['name']['first']} {self.form_data['name']['middle']}".strip()
            ws['B8'] = full_name
            ws['G8'] = self.form_data['date_filing'].strftime("%m/%d/%Y")
            ws['B9'] = self.form_data['position']
            ws['G9'] = self.form_data['salary']
            
            # Format inclusive dates
            start = self.form_data['inclusive_dates']['start']
            end = self.form_data['inclusive_dates']['end']
            if start == end:
                date_range = start.strftime("%B %d, %Y")
            else:
                date_range = f"{start.strftime('%B %d, %Y')} - {end.strftime('%B %d, %Y')}"
            ws['B12'] = date_range
            ws['G12'] = f"{self.form_data['working_days']} day" if self.form_data['working_days'] == 1 else f"{self.form_data['working_days']} days"
            
            # Save with timestamp
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"Leave_Application_{timestamp}.xlsx"
            wb.save(filename)
            wb.close()
            
            messagebox.showinfo("Success", f"Saved to {filename}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save Excel: {str(e)}")
    
    def generate_pdf(self):
        """Generate PDF with grid-based layout matching PhilHealth standard"""
        if not self.validate_inputs():
            return
        
        self.update_form_data()
        
        try:
            # Create output directory
            os.makedirs("output_pdf", exist_ok=True)
            os.makedirs("logos", exist_ok=True)
            
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"output_pdf/Application_for_Leave_{timestamp}.pdf"
            
            # Create PDF with grid system
            c = canvas.Canvas(filename, pagesize=A4)
            
            # Draw all sections using grid system
            self.draw_header_with_logos(c)
            self.draw_section_1(c)  # Office/Department & Name (now includes sections 1-5)
            self.draw_section_6(c)  # Details of Application
            self.draw_section_7(c)  # Details of Action
            
            c.save()
            messagebox.showinfo("Success", f"PDF generated: {filename}")
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to generate PDF: {str(e)}")
    
    def draw_header_with_logos(self, c):
        """Draw header section with proper PhilHealth format and logos"""
        # Header positioning
        header_y_start = CANVAS_HEIGHT - 20*mm
        
        # Stamp of Date of Receipt box (top right)
        stamp_x = CANVAS_WIDTH - MARGIN_RIGHT - 50*mm
        stamp_y = header_y_start - 10*mm
        c.rect(stamp_x, stamp_y, 45*mm, 20*mm)
        c.setFont("Helvetica", 8)
        c.drawString(stamp_x + 2*mm, stamp_y + 16*mm, "Stamp of Date of Receipt")
        
        # Civil Service Form header (top left)
        c.setFont("Helvetica", 8)
        c.drawString(MARGIN_LEFT, header_y_start - 5*mm, "Civil Service Form No. 6")
        c.drawString(MARGIN_LEFT, header_y_start - 10*mm, "Revised 2020")
        
        # Draw logos if available - FIXED positioning to avoid text overlap
        try:
            # PhilHealth logo (far left, below header text)
            if os.path.exists(self.philhealth_logo_var.get()):
                self.draw_logo(c, self.philhealth_logo_var.get(), 
                              MARGIN_LEFT, header_y_start - 70*mm, 35*mm, 25*mm)
            
            # Bagong Pilipinas logo (right of PhilHealth logo, below header text)
            if os.path.exists(self.bagong_logo_var.get()):
                self.draw_logo(c, self.bagong_logo_var.get(),
                              MARGIN_LEFT + 40*mm, header_y_start - 70*mm, 35*mm, 25*mm)
        except:
            pass  # Continue without logos if files not found
        
        # PhilHealth header text (centered) - FIXED positioning to avoid logo overlap
        y_pos = header_y_start - 30*mm  # Start text lower to avoid logo area
        
        c.setFont("Helvetica", 9)
        c.drawCentredString(CANVAS_WIDTH/2, y_pos, "Republic of the Philippines")
        y_pos -= 4*mm
        
        c.setFont("Helvetica-Bold", 10)
        c.drawCentredString(CANVAS_WIDTH/2, y_pos, "PHILIPPINE HEALTH INSURANCE CORPORATION")
        y_pos -= 4*mm
        
        c.setFont("Helvetica-Bold", 9)
        c.drawCentredString(CANVAS_WIDTH/2, y_pos, "PhilHealth Regional Office XI")
        y_pos -= 4*mm
        
        c.setFont("Helvetica", 8)
        c.drawCentredString(CANVAS_WIDTH/2, y_pos, "J.P. Laurel Avenue, Bajada, Poblacion District, Davao City")
        y_pos -= 3*mm
        c.drawCentredString(CANVAS_WIDTH/2, y_pos, "(082) 295-2133 local 6000, (082) 295-3385")
        y_pos -= 3*mm
        c.drawCentredString(CANVAS_WIDTH/2, y_pos, "teamphilhealth11 @ teamphilhealth @ www.philhealth.gov.ph")
        y_pos -= 10*mm  # Extra space before title
        
        # Main form title
        c.setFont("Helvetica-Bold", 14)
        c.drawCentredString(CANVAS_WIDTH/2, y_pos, "APPLICATION FOR LEAVE")
    
    def draw_logo(self, c, logo_path, x, y, max_width, max_height):
        """Draw logo with aspect ratio preservation"""
        try:
            img = Image.open(logo_path)
            img_width, img_height = img.size
            
            # Calculate scaling to fit within max dimensions
            scale_x = max_width / img_width
            scale_y = max_height / img_height
            scale = min(scale_x, scale_y)
            
            # Calculate final dimensions
            final_width = img_width * scale
            final_height = img_height * scale
            
            # Center the image within the allocated space
            final_x = x + (max_width - final_width) / 2
            final_y = y + (max_height - final_height) / 2
            
            c.drawImage(logo_path, final_x, final_y, final_width, final_height)
        except Exception as e:
            # Draw placeholder if logo fails
            c.rect(x, y, max_width, max_height)
            c.setFont("Helvetica", 8)
            c.drawString(x + 5, y + max_height/2, "Logo")
    
    def draw_section_1(self, c):
        """Draw Section 1-5 with proper table structure matching the reference image"""
        y_pos = SECTION_1_Y
        
        # Draw main border for sections 1-5
        table_width = CANVAS_WIDTH - 2*MARGIN_LEFT
        table_height = 30*mm
        c.rect(MARGIN_LEFT, y_pos - table_height, table_width, table_height)
        
        # Section 1 & 2 (first row)
        row1_y = y_pos - 5*mm
        c.setFont("Helvetica", 9)
        c.drawString(MARGIN_LEFT + 2*mm, row1_y, "1.   OFFICE/DEPARTMENT")
        c.drawString(MARGIN_LEFT + 90*mm, row1_y, "2.  NAME :")
        c.drawString(MARGIN_LEFT + 130*mm, row1_y, "(Last)")
        c.drawString(MARGIN_LEFT + 170*mm, row1_y, "(First)")
        c.drawString(MARGIN_LEFT + 210*mm, row1_y, "(Middle)")
        
        # Vertical dividers for first row
        c.line(MARGIN_LEFT + 85*mm, y_pos, MARGIN_LEFT + 85*mm, y_pos - 15*mm)  # After Office/Dept
        c.line(MARGIN_LEFT + 125*mm, y_pos - 5*mm, MARGIN_LEFT + 125*mm, y_pos - 15*mm)  # Before Last
        c.line(MARGIN_LEFT + 165*mm, y_pos - 5*mm, MARGIN_LEFT + 165*mm, y_pos - 15*mm)  # Before First
        c.line(MARGIN_LEFT + 205*mm, y_pos - 5*mm, MARGIN_LEFT + 205*mm, y_pos - 15*mm)  # Before Middle
        
        # Horizontal divider between rows
        c.line(MARGIN_LEFT, y_pos - 15*mm, MARGIN_LEFT + table_width, y_pos - 15*mm)
        
        # Data values for first row
        c.setFont("Helvetica", 10)
        c.drawString(MARGIN_LEFT + 2*mm, row1_y - 8*mm, self.form_data['office_department'])
        c.drawString(MARGIN_LEFT + 130*mm, row1_y - 8*mm, self.form_data['name']['last'])
        c.drawString(MARGIN_LEFT + 170*mm, row1_y - 8*mm, self.form_data['name']['first'])
        c.drawString(MARGIN_LEFT + 210*mm, row1_y - 8*mm, self.form_data['name']['middle'])
        
        # Section 3, 4, 5 (second row)
        row2_y = y_pos - 20*mm
        c.setFont("Helvetica", 9)
        c.drawString(MARGIN_LEFT + 2*mm, row2_y, "3.   DATE OF FILING")
        c.drawString(MARGIN_LEFT + 90*mm, row2_y, "4.   POSITION")
        c.drawString(MARGIN_LEFT + 170*mm, row2_y, "5.  SALARY")
        
        # Vertical dividers for second row
        c.line(MARGIN_LEFT + 85*mm, y_pos - 15*mm, MARGIN_LEFT + 85*mm, y_pos - table_height)
        c.line(MARGIN_LEFT + 165*mm, y_pos - 15*mm, MARGIN_LEFT + 165*mm, y_pos - table_height)
        
        # Data values for second row
        c.setFont("Helvetica", 10)
        date_str = self.form_data['date_filing'].strftime("%B %d, %Y")
        c.drawString(MARGIN_LEFT + 2*mm, row2_y - 5*mm, date_str)
        c.drawString(MARGIN_LEFT + 90*mm, row2_y - 5*mm, self.form_data['position'])
        c.drawString(MARGIN_LEFT + 170*mm, row2_y - 5*mm, self.form_data['salary'])
    

    def draw_section_6(self, c):
        """Draw Section 6 with proper table structure and borders like the reference image"""
        y_pos = SECTION_3_Y
        
        # Main section header
        c.setFont("Helvetica-Bold", 10)
        c.drawString(MARGIN_LEFT, y_pos, "6.  DETAILS OF APPLICATION")
        
        y_pos -= 10*mm
        
        # Draw main table border for section 6
        table_width = CANVAS_WIDTH - 2*MARGIN_LEFT
        table_height = 120*mm
        c.rect(MARGIN_LEFT, y_pos - table_height, table_width, table_height)
        
        # Column headers with border
        header_y = y_pos - 5*mm
        c.setFont("Helvetica-Bold", 9)
        c.drawString(MARGIN_LEFT + 2*mm, header_y, "6.A  TYPE OF LEAVE TO BE AVAILED OF")
        c.drawString(MARGIN_LEFT + 140*mm, header_y, "6.B  DETAILS OF LEAVE")
        
        # Vertical divider between columns
        col_divider_x = MARGIN_LEFT + 135*mm
        c.line(col_divider_x, y_pos, col_divider_x, y_pos - table_height)
        
        # Horizontal line under headers
        c.line(MARGIN_LEFT, y_pos - 10*mm, MARGIN_LEFT + table_width, y_pos - 10*mm)
        
        # Column A: Leave types with checkboxes
        leave_y = y_pos - 15*mm
        leave_types_text = [
            ('vacation', 'Vacation Leave (Sec. 51, Rule XVI, Omnibus Rules Implementing E.O. No. 292)'),
            ('mandatory', 'Mandatory/Forced Leave(Sec. 25, Rule XVI, Omnibus Rules Implementing E.O. No. 292)'),
            ('sick', 'Sick Leave  (Sec. 43, Rule XVI, Omnibus Rules Implementing E.O. No. 292)'),
            ('maternity', 'Maternity Leave (R.A. No. 11210 / IRR issued by CSC, DOLE and SSS)'),
            ('paternity', 'Paternity Leave (R.A. No. 8187 / CSC MC No. 71, s. 1998, as amended)'),
            ('special_privilege', 'Special Privilege Leave (Sec. 21, Rule XVI, Omnibus Rules Implementing E.O. No. 292)'),
            ('solo_parent', 'Solo Parent Leave (RA No. 8972 / CSC MC No. 8, s. 2004)'),
            ('study', 'Study Leave (Sec. 68, Rule XVI, Omnibus Rules Implementing E.O. No. 292)'),
            ('vawc_10day', '10-Day VAWC Leave (RA No. 9262 / CSC MC No. 15, s. 2005)'),
            ('rehabilitation', 'Rehabilitation Privilege (Sec. 55, Rule XVI, Omnibus Rules Implementing E.O. No. 292)'),
            ('special_women', 'Special Leave Benefits for Women (RA No. 9710 / CSC MC No. 25, s. 2010)'),
            ('special_emergency', 'Special Emergency (Calamity) Leave (CSC MC No. 2, s. 2012, as amended)'),
            ('adoption', 'Adoption Leave (R.A. No. 8552)')
        ]
        
        c.setFont("Helvetica", 7)
        for key, text in leave_types_text:
            # Draw checkbox
            is_checked = self.form_data['leave_types'].get(key, False)
            self.draw_checkbox(c, MARGIN_LEFT + 5*mm, leave_y - 1*mm, is_checked)
            
            # Draw text (truncated to fit column)
            max_width = 125*mm
            if c.stringWidth(text, "Helvetica", 7) > max_width:
                # Split long text into multiple lines
                words = text.split()
                line1 = ""
                line2 = ""
                for word in words:
                    if c.stringWidth(line1 + " " + word, "Helvetica", 7) < max_width:
                        line1 += (" " if line1 else "") + word
                    else:
                        line2 += (" " if line2 else "") + word
                
                c.drawString(MARGIN_LEFT + 10*mm, leave_y, line1)
                if line2:
                    leave_y -= 3*mm
                    c.drawString(MARGIN_LEFT + 10*mm, leave_y, line2)
            else:
                c.drawString(MARGIN_LEFT + 10*mm, leave_y, text)
            
            leave_y -= 6*mm
        
        # Others field
        c.drawString(MARGIN_LEFT + 5*mm, leave_y, "Others:")
        c.line(MARGIN_LEFT + 20*mm, leave_y - 1*mm, MARGIN_LEFT + 130*mm, leave_y - 1*mm)
        
        # Column B: Details of Leave
        details_y = y_pos - 15*mm
        c.setFont("Helvetica", 8)
        
        details_text = [
            "In case of Vacation/Special Privilege Leave:",
            "Within the Philippines",
            "_________________________________",
            "Abroad (Specify)",
            "_________________________________",
            "",
            "In case of Sick Leave:",
            "In Hospital (Specify Illness)",
            "_________________________________",
            "Out Patient (Specify Illness)",
            "_________________________________",
            "",
            "In case of Special Leave Benefits for Women:",
            "(Specify Illness)",
            "_________________________________",
            "",
            "In case of Study Leave:",
            "☐ Completion of Master's Degree",
            "☐ BAR/Board Examination Review",
            "",
            "Other purpose:",
            "☐ Monetization of Leave Credits",
            "☐ Terminal Leave"
        ]
        
        for line in details_text:
            if line.strip():
                if line.startswith("☐"):
                    # Draw checkbox for options
                    self.draw_checkbox(c, MARGIN_LEFT + 140*mm, details_y - 1*mm, False)
                    c.drawString(MARGIN_LEFT + 145*mm, details_y, line[2:])
                else:
                    c.drawString(MARGIN_LEFT + 140*mm, details_y, line)
                details_y -= 4*mm
            else:
                details_y -= 2*mm
        
        # Section 6.C and 6.D at bottom
        bottom_y = y_pos - table_height + 25*mm
        
        # Horizontal divider
        c.line(MARGIN_LEFT, bottom_y + 20*mm, MARGIN_LEFT + table_width, bottom_y + 20*mm)
        
        c.setFont("Helvetica-Bold", 9)
        c.drawString(MARGIN_LEFT + 2*mm, bottom_y + 15*mm, "6.C  NUMBER OF WORKING DAYS APPLIED FOR")
        c.drawString(MARGIN_LEFT + 140*mm, bottom_y + 15*mm, "6.D  COMMUTATION")
        
        # Working days value
        c.setFont("Helvetica-Bold", 10)
        working_days_text = f"{self.form_data['working_days']} Day" if self.form_data['working_days'] == 1 else f"{self.form_data['working_days']} Days"
        c.drawString(MARGIN_LEFT + 10*mm, bottom_y + 8*mm, working_days_text)
        
        # Commutation options
        c.setFont("Helvetica", 9)
        self.draw_checkbox(c, MARGIN_LEFT + 140*mm, bottom_y + 8*mm, self.form_data['commutation']['not_requested'])
        c.drawString(MARGIN_LEFT + 145*mm, bottom_y + 8*mm, "Not Requested")
        
        self.draw_checkbox(c, MARGIN_LEFT + 200*mm, bottom_y + 8*mm, self.form_data['commutation']['requested'])
        c.drawString(MARGIN_LEFT + 205*mm, bottom_y + 8*mm, "Requested")
        
        # Inclusive dates
        c.setFont("Helvetica", 9)
        c.drawString(MARGIN_LEFT + 10*mm, bottom_y, "INCLUSIVE DATES")
        
        start = self.form_data['inclusive_dates']['start']
        end = self.form_data['inclusive_dates']['end']
        if start == end:
            date_range = start.strftime("%B %d, %Y")
        else:
            date_range = f"{start.strftime('%B %d, %Y')} - {end.strftime('%B %d, %Y')}"
        
        c.setFont("Helvetica-Bold", 10)
        c.drawString(MARGIN_LEFT + 10*mm, bottom_y - 5*mm, date_range)
        
        # Signature line
        c.line(MARGIN_LEFT + 140*mm, bottom_y - 10*mm, MARGIN_LEFT + 220*mm, bottom_y - 10*mm)
        c.setFont("Helvetica", 8)
        c.drawString(MARGIN_LEFT + 165*mm, bottom_y - 15*mm, "(Signature of Applicant)")
    
    def draw_checkbox(self, c, x, y, is_checked):
        """Draw a physical vector checkbox (10px by 10px)"""
        box_size = 3*mm
        c.rect(x, y, box_size, box_size)
        
        if is_checked:
            # Draw checkmark
            c.line(x + 0.5*mm, y + 1.5*mm, x + 1.2*mm, y + 0.8*mm)
            c.line(x + 1.2*mm, y + 0.8*mm, x + 2.5*mm, y + 2.2*mm)
    
    def wrap_text(self, text, max_chars):
        """Wrap text to specified character width"""
        return textwrap.wrap(text, width=max_chars)
    
    def truncate_text(self, text, max_chars):
        """Truncate text with ellipsis if too long"""
        if len(text) <= max_chars:
            return text
        return text[:max_chars-3] + "..."
    
    def draw_section_7(self, c):
        """Draw Section 7: Details of Action on Application with proper table structure"""
        y_pos = SECTION_7_Y
        
        # Section header
        c.setFont("Helvetica-Bold", 10)
        c.drawString(MARGIN_LEFT, y_pos, "7.  DETAILS OF ACTION ON APPLICATION")
        
        y_pos -= 10*mm
        
        # Draw main table border for section 7
        table_width = CANVAS_WIDTH - 2*MARGIN_LEFT
        table_height = 80*mm
        c.rect(MARGIN_LEFT, y_pos - table_height, table_width, table_height)
        
        # 7.A and 7.B headers
        header_y = y_pos - 5*mm
        c.setFont("Helvetica-Bold", 9)
        c.drawString(MARGIN_LEFT + 2*mm, header_y, "7.A  CERTIFICATION OF LEAVE CREDITS")
        c.drawString(MARGIN_LEFT + 140*mm, header_y, "7.B  RECOMMENDATION")
        
        # Vertical divider
        col_divider_x = MARGIN_LEFT + 135*mm
        c.line(col_divider_x, y_pos, col_divider_x, y_pos - table_height)
        
        # Horizontal line under headers
        c.line(MARGIN_LEFT, y_pos - 10*mm, MARGIN_LEFT + table_width, y_pos - 10*mm)
        
        # 7.A Content
        content_y = y_pos - 15*mm
        c.setFont("Helvetica", 8)
        c.drawString(MARGIN_LEFT + 10*mm, content_y, "As of")
        c.line(MARGIN_LEFT + 20*mm, content_y - 1*mm, MARGIN_LEFT + 60*mm, content_y - 1*mm)
        
        content_y -= 8*mm
        
        # Leave credits table
        table_x = MARGIN_LEFT + 10*mm
        table_y = content_y
        
        # Table headers
        c.rect(table_x, table_y - 15*mm, 80*mm, 15*mm)
        c.line(table_x + 40*mm, table_y, table_x + 40*mm, table_y - 15*mm)  # Vertical divider
        c.line(table_x, table_y - 5*mm, table_x + 80*mm, table_y - 5*mm)   # Horizontal divider
        
        c.setFont("Helvetica", 8)
        c.drawString(table_x + 15*mm, table_y - 3*mm, "Vacation Leave")
        c.drawString(table_x + 55*mm, table_y - 3*mm, "Sick Leave")
        
        c.drawString(table_x + 2*mm, table_y - 8*mm, "Total Earned")
        c.drawString(table_x + 2*mm, table_y - 11*mm, "Less this application")
        c.drawString(table_x + 2*mm, table_y - 14*mm, "Balance")
        
        # Authorized Officer signature line
        sig_y = table_y - 25*mm
        c.line(table_x, sig_y, table_x + 80*mm, sig_y)
        c.setFont("Helvetica", 7)
        c.drawString(table_x + 25*mm, sig_y - 4*mm, "(Authorized Officer)")
        
        # 7.B Content
        c.setFont("Helvetica", 8)
        rec_y = y_pos - 15*mm
        self.draw_checkbox(c, MARGIN_LEFT + 140*mm, rec_y - 1*mm, False)
        c.drawString(MARGIN_LEFT + 145*mm, rec_y, "For approval")
        
        rec_y -= 5*mm
        self.draw_checkbox(c, MARGIN_LEFT + 140*mm, rec_y - 1*mm, False)
        c.drawString(MARGIN_LEFT + 145*mm, rec_y, "For disapproval due to")
        c.line(MARGIN_LEFT + 185*mm, rec_y - 1*mm, MARGIN_LEFT + 240*mm, rec_y - 1*mm)
        
        # Additional lines for disapproval reasons
        for i in range(3):
            rec_y -= 4*mm
            c.line(MARGIN_LEFT + 140*mm, rec_y - 1*mm, MARGIN_LEFT + 240*mm, rec_y - 1*mm)
        
        # Authorized Officer signature for 7.B
        rec_y -= 10*mm
        c.line(MARGIN_LEFT + 160*mm, rec_y, MARGIN_LEFT + 220*mm, rec_y)
        c.setFont("Helvetica", 7)
        c.drawString(MARGIN_LEFT + 175*mm, rec_y - 4*mm, "(Authorized Officer)")
        
        # 7.C and 7.D at bottom
        bottom_y = y_pos - table_height + 30*mm
        
        # Horizontal divider
        c.line(MARGIN_LEFT, bottom_y + 25*mm, MARGIN_LEFT + table_width, bottom_y + 25*mm)
        
        c.setFont("Helvetica-Bold", 9)
        c.drawString(MARGIN_LEFT + 2*mm, bottom_y + 20*mm, "7.C  APPROVED FOR:")
        c.drawString(MARGIN_LEFT + 140*mm, bottom_y + 20*mm, "7.D   DISAPPROVED DUE TO:")
        
        # 7.C Content
        c.setFont("Helvetica", 8)
        c.line(MARGIN_LEFT + 10*mm, bottom_y + 15*mm, MARGIN_LEFT + 25*mm, bottom_y + 15*mm)
        c.drawString(MARGIN_LEFT + 27*mm, bottom_y + 15*mm, "days with pay")
        
        c.line(MARGIN_LEFT + 10*mm, bottom_y + 10*mm, MARGIN_LEFT + 25*mm, bottom_y + 10*mm)
        c.drawString(MARGIN_LEFT + 27*mm, bottom_y + 10*mm, "days without pay")
        
        c.line(MARGIN_LEFT + 10*mm, bottom_y + 5*mm, MARGIN_LEFT + 25*mm, bottom_y + 5*mm)
        c.drawString(MARGIN_LEFT + 27*mm, bottom_y + 5*mm, "others (Specify)")
        
        # 7.D Content (lines for disapproval reasons)
        for i in range(3):
            line_y = bottom_y + 15*mm - (i * 5*mm)
            c.line(MARGIN_LEFT + 140*mm, line_y, MARGIN_LEFT + 240*mm, line_y)
        
        # Final authorized official signature
        final_sig_y = bottom_y - 10*mm
        c.line(MARGIN_LEFT + 80*mm, final_sig_y, MARGIN_LEFT + 160*mm, final_sig_y)
        c.setFont("Helvetica", 7)
        c.drawString(MARGIN_LEFT + 105*mm, final_sig_y - 4*mm, "(Authorized Official)")
    
    def clear_form(self):
        """Clear all form fields except Office/Department default"""
        self.name_last_var.set("")
        self.name_first_var.set("")
        self.name_middle_var.set("")
        self.position_var.set("")
        self.salary_var.set("")
        self.working_days_var.set("")
        
        # Clear leave type checkboxes
        for var in self.leave_vars.values():
            var.set(False)
        
        # Reset dates
        self.date_filing.set_date(datetime.now())
        self.start_date.set_date(datetime.now())
        self.end_date.set_date(datetime.now())
        
        # Reset commutation
        self.comm_var.set("not_requested")
        
        # Keep Office/Department as default

def main():
    root = tk.Tk()
    app = LeaveApplicationEncoder(root)
    root.mainloop()

if __name__ == "__main__":
    main()