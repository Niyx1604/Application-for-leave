#!/usr/bin/env python3
"""
Leave Application Encoder - Civil Service Form No. 6 (Revised 2020)
Preserves exact Excel structure and generates identical PDF output
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from tkcalendar import DateEntry
from datetime import datetime, date
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib.units import mm
import os

class LeaveApplicationEncoder:
    def __init__(self, root):
        self.root = root
        self.root.title("Leave Application Encoder - Civil Service Form No. 6")
        self.root.geometry("700x600")
        
        # Position options with roman numerals preserved
        self.positions = [
            "Administrative Aide I",
            "Administrative Aide II",
            "Administrative Aide III",
            "Administrative Aide IV",
            "Administrative Aide V",
            "Administrative Aide VI",
            "Social Insurance Assistant I",
            "Social Insurance Assistant II",
            "Social Insurance Assistant III",
            "Social Insurance Assistant IV",
            "Social Insurance Assistant V",
            "Social Insurance Officer I",
            "Social Insurance Officer II",
            "Social Insurance Officer III",
            "Chief Social Insurance Officer I"
        ]
        
        self.create_widgets()
        
    def create_widgets(self):
        # Main frame
        main_frame = ttk.Frame(self.root, padding="20")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Title section - preserving exact Excel text
        title_frame = ttk.LabelFrame(main_frame, text="Form Header", padding="10")
        title_frame.grid(row=0, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(0, 15))
        
        ttk.Label(title_frame, text="CIVIL SERVICE FORM NO. 6", 
                 font=("Arial", 12, "bold")).grid(row=0, column=0, pady=2)
        ttk.Label(title_frame, text="(Revised 2020)").grid(row=1, column=0)
        ttk.Label(title_frame, text="APPLICATION FOR LEAVE", 
                 font=("Arial", 11, "bold")).grid(row=2, column=0, pady=2)
        
        # Office/Department - preserving default value "LHIO - Digos"
        ttk.Label(main_frame, text="Office/Department:").grid(row=1, column=0, sticky=tk.W, pady=5)
        self.office_var = tk.StringVar(value="LHIO - Digos")
        ttk.Entry(main_frame, textvariable=self.office_var, width=40).grid(
            row=1, column=1, sticky=(tk.W, tk.E), pady=5)
        
        # Name field
        ttk.Label(main_frame, text="Name:").grid(row=2, column=0, sticky=tk.W, pady=5)
        self.name_var = tk.StringVar()
        ttk.Entry(main_frame, textvariable=self.name_var, width=40).grid(
            row=2, column=1, sticky=(tk.W, tk.E), pady=5)
        
        # Date of Filing - auto-populated with today's date
        ttk.Label(main_frame, text="Date of Filing:").grid(row=3, column=0, sticky=tk.W, pady=5)
        self.date_filing = DateEntry(main_frame, width=37, background='darkblue',
                                     foreground='white', borderwidth=2, 
                                     date_pattern='mm/dd/yyyy')
        self.date_filing.grid(row=3, column=1, sticky=(tk.W, tk.E), pady=5)
        
        # Position dropdown with exact options
        ttk.Label(main_frame, text="Position:").grid(row=4, column=0, sticky=tk.W, pady=5)
        self.position_var = tk.StringVar()
        position_combo = ttk.Combobox(main_frame, textvariable=self.position_var,
                                     values=self.positions, state="readonly", width=37)
        position_combo.grid(row=4, column=1, sticky=(tk.W, tk.E), pady=5)
        
        # Salary field
        ttk.Label(main_frame, text="Salary:").grid(row=5, column=0, sticky=tk.W, pady=5)
        self.salary_var = tk.StringVar()
        ttk.Entry(main_frame, textvariable=self.salary_var, width=40).grid(
            row=5, column=1, sticky=(tk.W, tk.E), pady=5)
        
        # Inclusive Dates section
        dates_frame = ttk.LabelFrame(main_frame, text="Inclusive Dates", padding="10")
        dates_frame.grid(row=6, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=10)
        
        ttk.Label(dates_frame, text="Start Date:").grid(row=0, column=0, sticky=tk.W, pady=5)
        self.start_date = DateEntry(dates_frame, width=18, background='darkblue',
                                    foreground='white', borderwidth=2, 
                                    date_pattern='mm/dd/yyyy')
        self.start_date.grid(row=0, column=1, sticky=(tk.W, tk.E), pady=5, padx=(5, 0))
        
        ttk.Label(dates_frame, text="End Date:").grid(row=1, column=0, sticky=tk.W, pady=5)
        self.end_date = DateEntry(dates_frame, width=18, background='darkblue',
                                  foreground='white', borderwidth=2,
                                  date_pattern='mm/dd/yyyy')
        self.end_date.grid(row=1, column=1, sticky=(tk.W, tk.E), pady=5, padx=(5, 0))
        
        # Working Days dropdown (1-31 with "day/days" format)
        ttk.Label(main_frame, text="Working Days:").grid(row=7, column=0, sticky=tk.W, pady=5)
        working_days_options = [f"{i} day" if i == 1 else f"{i} days" for i in range(1, 32)]
        self.working_days_var = tk.StringVar()
        working_days_combo = ttk.Combobox(main_frame, textvariable=self.working_days_var,
                                         values=working_days_options, state="readonly", width=37)
        working_days_combo.grid(row=7, column=1, sticky=(tk.W, tk.E), pady=5)
        
        # Buttons frame
        button_frame = ttk.Frame(main_frame)
        button_frame.grid(row=8, column=0, columnspan=2, pady=20)
        
        ttk.Button(button_frame, text="Generate PDF", 
                  command=self.generate_pdf).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Save to Excel", 
                  command=self.save_to_excel).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Clear Form", 
                  command=self.clear_form).pack(side=tk.LEFT, padx=5)
        
        # Configure grid weights
        main_frame.columnconfigure(1, weight=1)
        dates_frame.columnconfigure(1, weight=1)
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
    
    def validate_inputs(self):
        """Validate required fields"""
        if not self.name_var.get().strip():
            messagebox.showerror("Validation Error", "Name is required")
            return False
        if not self.position_var.get():
            messagebox.showerror("Validation Error", "Position must be selected")
            return False
        if not self.working_days_var.get():
            messagebox.showerror("Validation Error", "Working Days must be selected")
            return False
        return True
    
    def save_to_excel(self):
        """Save data to Excel file preserving exact structure"""
        if not self.validate_inputs():
            return
        
        try:
            # Load the template
            wb = openpyxl.load_workbook('ALA.xlsx')
            ws = wb.active
            
            # Populate cells - preserving exact cell locations
            ws['B7'] = self.office_var.get()
            ws['B8'] = self.name_var.get()
            ws['G8'] = self.date_filing.get_date().strftime("%m/%d/%Y")
            ws['B9'] = self.position_var.get()
            ws['G9'] = self.salary_var.get()
            
            # Format inclusive dates
            start = self.start_date.get_date()
            end = self.end_date.get_date()
            if start == end:
                date_range = start.strftime("%B %d, %Y")
            else:
                date_range = f"{start.strftime('%B %d, %Y')} - {end.strftime('%B %d, %Y')}"
            ws['B12'] = date_range
            ws['G12'] = self.working_days_var.get()
            
            # Save with timestamp
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"Leave_Application_{timestamp}.xlsx"
            wb.save(filename)
            wb.close()
            
            messagebox.showinfo("Success", f"Saved to {filename}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save Excel: {str(e)}")
    
    def generate_pdf(self):
        """Generate PDF with identical layout to Excel"""
        if not self.validate_inputs():
            return
        
        try:
            # Create output directory
            os.makedirs("output_pdf", exist_ok=True)
            
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"output_pdf/Application_for_Leave_{timestamp}.pdf"
            
            # Create PDF
            c = canvas.Canvas(filename, pagesize=A4)
            width, height = A4
            
            # Set font
            c.setFont("Helvetica-Bold", 12)
            
            # Title - preserving exact text
            c.drawCentredString(width/2, height - 50*mm, "CIVIL SERVICE FORM NO. 6")
            c.setFont("Helvetica", 10)
            c.drawCentredString(width/2, height - 55*mm, "(Revised 2020)")
            c.setFont("Helvetica-Bold", 11)
            c.drawCentredString(width/2, height - 62*mm, "APPLICATION FOR LEAVE")
            
            # Form fields - preserving exact labels and layout
            y_pos = height - 80*mm
            c.setFont("Helvetica", 10)
            
            # Office/Department
            c.drawString(30*mm, y_pos, "Office/Department:")
            c.drawString(70*mm, y_pos, self.office_var.get())
            y_pos -= 10*mm
            
            # Name and Date of Filing on same line
            c.drawString(30*mm, y_pos, "Name:")
            c.drawString(50*mm, y_pos, self.name_var.get())
            c.drawString(130*mm, y_pos, "Date of Filing:")
            c.drawString(160*mm, y_pos, self.date_filing.get_date().strftime("%m/%d/%Y"))
            y_pos -= 10*mm
            
            # Position and Salary on same line
            c.drawString(30*mm, y_pos, "Position:")
            c.drawString(50*mm, y_pos, self.position_var.get())
            c.drawString(130*mm, y_pos, "Salary:")
            c.drawString(150*mm, y_pos, self.salary_var.get())
            y_pos -= 15*mm
            
            # Inclusive Dates
            start = self.start_date.get_date()
            end = self.end_date.get_date()
            if start == end:
                date_range = start.strftime("%B %d, %Y")
            else:
                date_range = f"{start.strftime('%B %d, %Y')} - {end.strftime('%B %d, %Y')}"
            
            c.drawString(30*mm, y_pos, "Inclusive Dates:")
            c.drawString(60*mm, y_pos, date_range)
            c.drawString(130*mm, y_pos, "Working Days:")
            c.drawString(160*mm, y_pos, self.working_days_var.get())
            
            # Save PDF
            c.save()
            
            messagebox.showinfo("Success", f"PDF generated: {filename}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to generate PDF: {str(e)}")
    
    def clear_form(self):
        """Clear all form fields except Office/Department default"""
        self.name_var.set("")
        self.position_var.set("")
        self.salary_var.set("")
        self.working_days_var.set("")
        self.date_filing.set_date(datetime.now())
        self.start_date.set_date(datetime.now())
        self.end_date.set_date(datetime.now())
        # Keep Office/Department as "LHIO - Digos"

def main():
    root = tk.Tk()
    app = LeaveApplicationEncoder(root)
    root.mainloop()

if __name__ == "__main__":
    main()
